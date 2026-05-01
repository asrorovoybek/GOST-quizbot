import asyncio
import json
import os
import random
import time
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set

import asyncpg
import openpyxl
from io import BytesIO
from aiohttp import web
from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ChatType, ParseMode, PollType
from aiogram.filters import Command, CommandStart
from aiogram.types import (
    BotCommand,
    BotCommandScopeAllPrivateChats,
    BotCommandScopeChat,
    BufferedInputFile,
    CallbackQuery,
    Document,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    PollAnswer,
    ReplyKeyboardMarkup,
)

# ============================================================
# 1) BOT TOKEN VA ADMINLAR
# ============================================================
BOT_TOKEN = os.getenv("BOT_TOKEN")
DATABASE_URL = os.getenv("DATABASE_URL")
# Admin ID larini environment variable dan olamiz
raw_admin_ids = os.getenv("ADMIN_IDS", "5383208910")
ADMIN_IDS = {int(x.strip()) for x in raw_admin_ids.split(",") if x.strip().isdigit()}
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "@GlobalJournals_site")  # admin username
PAYMENT_INFO = os.getenv("PAYMENT_INFO", "Admin bilan bog'laning.")  # to'lov ma'lumotlari
ADMIN_CONTACT_ID = 8566281882  # admin bilan bog'lanish uchun Telegram ID

# ============================================================
# 2) SOZLAMALAR
# ============================================================
POLL_OPEN_PERIOD = 60
MIN_OPTIONS = 2
MAX_OPTIONS = 10
QUESTIONS_PER_SESSION = 50
FREE_QUESTIONS_LIMIT = 5  # Bepul savollar soni

# ============================================================
# 3) TEST YUKLASH SHABLONI
# ============================================================
TEMPLATE_TEXT = """#subject: Fan nomi
#short: Qisqa nom
#question: Savol matni 1
* To'g'ri javob
- Noto'g'ri javob 1
- Noto'g'ri javob 2
- Noto'g'ri javob 3

#question: Savol matni 2
- Variant A
* Variant B
- Variant C
- Variant D
"""

# ============================================================
# 4) DATABASE POOL
# ============================================================
db_pool: Optional[asyncpg.Pool] = None


async def get_pool() -> asyncpg.Pool:
    global db_pool
    if db_pool is None:
        db_pool = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=5)
    return db_pool


async def init_db():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                user_id BIGINT PRIMARY KEY,
                full_name TEXT DEFAULT 'Foydalanuvchi',
                username TEXT DEFAULT '',
                is_approved BOOLEAN DEFAULT FALSE,
                is_blocked BOOLEAN DEFAULT FALSE,
                free_questions_used INT DEFAULT 0
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS user_subjects (
                id SERIAL PRIMARY KEY,
                user_id BIGINT REFERENCES users(user_id) ON DELETE CASCADE,
                subject TEXT NOT NULL,
                used_question_ids JSONB DEFAULT '[]',
                attempts INT DEFAULT 0,
                best_score INT DEFAULT 0,
                best_total INT DEFAULT 0,
                best_percent FLOAT DEFAULT 0,
                best_time_seconds INT DEFAULT 0,
                best_achieved_at BIGINT DEFAULT 0,
                last_score INT DEFAULT 0,
                last_total INT DEFAULT 0,
                last_percent FLOAT DEFAULT 0,
                last_time_seconds INT DEFAULT 0,
                UNIQUE(user_id, subject)
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS subjects (
                subject TEXT PRIMARY KEY,
                short_name TEXT DEFAULT '',
                questions JSONB NOT NULL DEFAULT '[]'
            )
        """)


# ============================================================
# 5) DATA CLASSLAR
# ============================================================
@dataclass
class Question:
    question: str
    options: List[str]
    correct_option_id: int

    def to_dict(self) -> dict:
        return {
            "question": self.question,
            "options": self.options,
            "correct_option_id": self.correct_option_id,
        }

    @staticmethod
    def from_dict(data: dict) -> "Question":
        return Question(
            question=data["question"],
            options=data["options"],
            correct_option_id=data["correct_option_id"],
        )


@dataclass
class UserSession:
    subject: Optional[str] = None
    question_ids: List[int] = field(default_factory=list)
    current_index: int = 0
    score: int = 0
    waiting_poll: bool = False
    current_question_id: Optional[int] = None
    current_poll_id: Optional[str] = None
    current_poll_message_id: Optional[int] = None
    current_control_message_id: Optional[int] = None
    started_at: Optional[float] = None
    poll_sent_at: Optional[float] = None
    paused: bool = False


router = Router()
user_sessions: Dict[int, UserSession] = {}
poll_to_user: Dict[str, int] = {}
subject_tests: Dict[str, List[Question]] = {}
subject_short_names: Dict[str, str] = {}  # {to'liq nom: qisqa nom}
admin_broadcast_state: Dict[int, bool] = {}  # broadcast rejimi

# ============================================================
# 6) DATABASE FUNKSIYALARI
# ============================================================
async def load_questions() -> tuple[Dict[str, List[Question]], Dict[str, str]]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT subject, short_name, questions FROM subjects")
    loaded: Dict[str, List[Question]] = {}
    shorts: Dict[str, str] = {}
    for row in rows:
        questions_raw = row["questions"]
        if isinstance(questions_raw, str):
            questions_raw = json.loads(questions_raw)
        loaded[row["subject"]] = [Question.from_dict(q) for q in questions_raw]
        shorts[row["subject"]] = row["short_name"] or ""
    return loaded, shorts


async def save_subject(subject: str, questions: List[Question], short_name: str = "") -> None:
    pool = await get_pool()
    questions_json = json.dumps([q.to_dict() for q in questions], ensure_ascii=False)
    async with pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO subjects (subject, short_name, questions)
            VALUES ($1, $2, $3::jsonb)
            ON CONFLICT (subject) DO UPDATE SET
                short_name = EXCLUDED.short_name,
                questions = EXCLUDED.questions
        """, subject, short_name, questions_json)


async def delete_subject_db(subject: str) -> None:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM subjects WHERE subject = $1", subject)


async def delete_user_db(user_id: int) -> None:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM users WHERE user_id = $1", user_id)


async def ensure_user_record_db(
    user_id: int,
    full_name: Optional[str] = None,
    username: Optional[str] = None,
) -> dict:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM users WHERE user_id = $1", user_id)
        if row is None:
            await conn.execute("""
                INSERT INTO users (user_id, full_name, username, is_approved, is_blocked)
                VALUES ($1, $2, $3, FALSE, FALSE)
            """, user_id, full_name or "Foydalanuvchi", username or "")
            row = await conn.fetchrow("SELECT * FROM users WHERE user_id = $1", user_id)
        else:
            updates = []
            params = []
            idx = 1
            if full_name:
                updates.append(f"full_name = ${idx}")
                params.append(full_name)
                idx += 1
            if username is not None:
                updates.append(f"username = ${idx}")
                params.append(username)
                idx += 1
            if updates:
                params.append(user_id)
                await conn.execute(
                    f"UPDATE users SET {', '.join(updates)} WHERE user_id = ${idx}",
                    *params,
                )
            row = await conn.fetchrow("SELECT * FROM users WHERE user_id = $1", user_id)
    return dict(row)



async def get_free_questions_used(user_id: int) -> int:
    user = await get_user_record(user_id)
    if not user:
        return 0
    return user.get("free_questions_used", 0) or 0


async def increment_free_questions(user_id: int) -> int:
    """Bepul savollar sonini +1 qiladi va yangi qiymatni qaytaradi."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        # Agar ustun mavjud bo'lmasa, avval qo'shib olamiz
        try:
            await conn.execute(
                "ALTER TABLE users ADD COLUMN IF NOT EXISTS free_questions_used INT DEFAULT 0"
            )
        except Exception:
            pass
        row = await conn.fetchrow(
            "UPDATE users SET free_questions_used = COALESCE(free_questions_used, 0) + 1 "
            "WHERE user_id = $1 RETURNING free_questions_used",
            user_id
        )
    return row["free_questions_used"] if row else 1


async def has_free_limit_reached(user_id: int) -> bool:
    """Foydalanuvchi bepul limitga yetganmi?"""
    used = await get_free_questions_used(user_id)
    return used >= FREE_QUESTIONS_LIMIT

async def get_user_record(user_id: int) -> Optional[dict]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM users WHERE user_id = $1", user_id)
    return dict(row) if row else None


async def set_user_approved(user_id: int, approved: bool, blocked: bool = False) -> None:
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE users SET is_approved = $1, is_blocked = $2 WHERE user_id = $3",
            approved, blocked, user_id
        )


async def set_user_blocked(user_id: int, blocked: bool) -> None:
    pool = await get_pool()
    async with pool.acquire() as conn:
        if blocked:
            await conn.execute(
                "UPDATE users SET is_blocked = TRUE, is_approved = FALSE WHERE user_id = $1",
                user_id
            )
        else:
            await conn.execute(
                "UPDATE users SET is_blocked = FALSE WHERE user_id = $1",
                user_id
            )


async def ensure_user_subject_record_db(
    user_id: int,
    subject: str,
    full_name: Optional[str] = None,
    username: Optional[str] = None,
) -> dict:
    await ensure_user_record_db(user_id, full_name=full_name, username=username)
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT * FROM user_subjects WHERE user_id = $1 AND subject = $2",
            user_id, subject
        )
        if row is None:
            await conn.execute("""
                INSERT INTO user_subjects (user_id, subject)
                VALUES ($1, $2)
                ON CONFLICT (user_id, subject) DO NOTHING
            """, user_id, subject)
            row = await conn.fetchrow(
                "SELECT * FROM user_subjects WHERE user_id = $1 AND subject = $2",
                user_id, subject
            )
    return dict(row)


async def save_user_subject(user_id: int, subject: str, data: dict) -> None:
    pool = await get_pool()
    used_ids_json = json.dumps(data.get("used_question_ids", []))
    async with pool.acquire() as conn:
        await conn.execute("""
            UPDATE user_subjects SET
                used_question_ids = $1::jsonb,
                attempts = $2,
                best_score = $3,
                best_total = $4,
                best_percent = $5,
                best_time_seconds = $6,
                best_achieved_at = $7,
                last_score = $8,
                last_total = $9,
                last_percent = $10,
                last_time_seconds = $11
            WHERE user_id = $12 AND subject = $13
        """,
            used_ids_json,
            data.get("attempts", 0),
            data.get("best_score", 0),
            data.get("best_total", 0),
            data.get("best_percent", 0.0),
            data.get("best_time_seconds", 0),
            data.get("best_achieved_at", 0),
            data.get("last_score", 0),
            data.get("last_total", 0),
            data.get("last_percent", 0.0),
            data.get("last_time_seconds", 0),
            user_id, subject
        )


async def is_user_approved(user_id: int) -> bool:
    user = await get_user_record(user_id)
    if not user:
        return False
    return user.get("is_approved", False) and not user.get("is_blocked", False)


async def is_user_blocked(user_id: int) -> bool:
    user = await get_user_record(user_id)
    if not user:
        return False
    return user.get("is_blocked", False)


async def get_pending_users() -> List[dict]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT * FROM users WHERE is_approved = FALSE AND is_blocked = FALSE"
        )
    return [dict(r) for r in rows]


async def get_all_users_with_subjects() -> List[dict]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT u.user_id, u.full_name, u.username,
                   s.subject, s.best_percent, s.best_score, s.best_total,
                   s.best_time_seconds, s.best_achieved_at
            FROM users u
            JOIN user_subjects s ON u.user_id = s.user_id
        """)
    return [dict(r) for r in rows]


async def get_user_all_subjects(user_id: int) -> List[dict]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT * FROM user_subjects WHERE user_id = $1", user_id
        )
    return [dict(r) for r in rows]


async def get_all_approved_user_ids() -> List[int]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT user_id FROM users WHERE is_approved = TRUE AND is_blocked = FALSE"
        )
    return [r["user_id"] for r in rows]


async def get_all_approved_users() -> List[dict]:
    pool = await get_pool()
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT user_id, full_name, username FROM users WHERE is_approved = TRUE AND is_blocked = FALSE ORDER BY user_id"
        )
    return [dict(r) for r in rows]


# ============================================================
# 7) YORDAMCHI FUNKSIYALAR
# ============================================================
def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def format_duration(seconds: int) -> str:
    minutes = seconds // 60
    sec = seconds % 60
    return f"{minutes}m {sec}s"


def get_main_reply_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📚 Fanlar"), KeyboardButton(text="🏆 Reyting")],
            [KeyboardButton(text="📊 Mening natijam"), KeyboardButton(text="ℹ️ Yordam")],
        ],
        resize_keyboard=True,
        is_persistent=True,
        input_field_placeholder="Fan tanlang yoki buyruq yuboring...",
    )


def get_display_name(subject: str) -> str:
    """Tugmada ko'rsatiladigan nom: qisqa nom bo'lsa uni, bo'lmasa to'liq nomni qisqartiradi."""
    short = subject_short_names.get(subject, "")
    if short:
        return short
    if len(subject) <= 25:
        return subject
    return subject[:24] + "…"


def build_subjects_text_and_keyboard() -> tuple[str, Optional[InlineKeyboardMarkup]]:
    """
    Fanlar uchun matn (to'liq nomlar) va klaviatura (qisqa nomlar) qaytaradi.
    """
    subjects = sorted(subject_tests.keys())
    if not subjects:
        return "", None

    # Yuqorida to'liq nomlar ro'yxati
    lines = ["📚 <b>Mavjud fanlar:</b>\n"]
    for idx, subject in enumerate(subjects, start=1):
        short = subject_short_names.get(subject, "")
        if short and short != subject:
            lines.append(f"{idx}. {subject} <i>({short})</i>")
        else:
            lines.append(f"{idx}. {subject}")
    text = "\n".join(lines)

    # Tugmalarda qisqa nom yoki to'liq nom
    rows = []
    for idx, subject in enumerate(subjects):
        display = get_display_name(subject)
        rows.append(
            [InlineKeyboardButton(text=f"{idx+1}. 📘 {display}", callback_data=f"subj:{idx}")]
        )
    rows.append([InlineKeyboardButton(text="🔄 Yangilash", callback_data="refresh_subjects")])

    return text, InlineKeyboardMarkup(inline_keyboard=rows)


def build_subjects_inline_keyboard() -> Optional[InlineKeyboardMarkup]:
    _, keyboard = build_subjects_text_and_keyboard()
    return keyboard


def build_finish_test_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="⏹ Testni tugatish", callback_data="finish_test")]
        ]
    )


def build_paused_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="▶️ Testni davom ettirish", callback_data="continue_test")],
            [InlineKeyboardButton(text="⏹ Testni tugatish", callback_data="finish_test")],
        ]
    )



def build_payment_keyboard() -> InlineKeyboardMarkup:
    """To'lov so'rovi — kvitansiya yuborish + adminga murojaat."""
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📎 Kvitansiya yuborish", callback_data="send_receipt")],
            [InlineKeyboardButton(text="💬 Admin bilan bog'lanish", url=f"https://t.me/{ADMIN_USERNAME.lstrip('@')}")],
        ]
    )

def build_access_request_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(
                text="💬 Admin bilan bog'lanish",
                url=f"https://t.me/GlobalJournals_site"
            )],
            [InlineKeyboardButton(text="🔐 Ruxsat so'rash", callback_data="request_access")],
        ]
    )


def build_admin_approval_keyboard(user_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Ruxsat berish", callback_data=f"approve_user:{user_id}"),
                InlineKeyboardButton(text="❌ Rad etish", callback_data=f"reject_user:{user_id}"),
            ],
            [
                InlineKeyboardButton(text="⛔ Bloklash", callback_data=f"block_user:{user_id}")
            ],
        ]
    )


def get_or_create_session(user_id: int) -> UserSession:
    if user_id not in user_sessions:
        user_sessions[user_id] = UserSession()
    return user_sessions[user_id]


def parse_subject_file(text: str) -> tuple[str, str, List[Question]]:
    """(subject, short_name, questions) qaytaradi."""
    lines = [line.strip() for line in text.splitlines()]
    subject: Optional[str] = None
    short_name: str = ""
    questions: List[Question] = []

    current_question: Optional[str] = None
    current_options: List[str] = []
    correct_index: Optional[int] = None

    def finalize_question() -> None:
        nonlocal current_question, current_options, correct_index
        if current_question is None:
            return
        if len(current_options) < MIN_OPTIONS:
            raise ValueError(
                f"'{current_question}' savolida variantlar kamida {MIN_OPTIONS} ta bo'lishi kerak."
            )
        if len(current_options) > MAX_OPTIONS:
            raise ValueError(
                f"'{current_question}' savolida variantlar ko'pi bilan {MAX_OPTIONS} ta bo'lishi kerak."
            )
        if correct_index is None:
            raise ValueError(
                f"'{current_question}' savolida to'g'ri javob * bilan belgilanmagan."
            )
        paired = list(enumerate(current_options))
        random.shuffle(paired)
        shuffled_options = [option for _, option in paired]
        shuffled_correct_index = next(
            new_index for new_index, (old_index, _) in enumerate(paired)
            if old_index == correct_index
        )
        questions.append(
            Question(
                question=current_question,
                options=shuffled_options,
                correct_option_id=shuffled_correct_index,
            )
        )
        current_question = None
        current_options = []
        correct_index = None

    for line in lines:
        if not line:
            continue
        if line.startswith("#subject:"):
            subject = line.split(":", 1)[1].strip()
            continue
        if line.startswith("#short:"):
            short_name = line.split(":", 1)[1].strip()
            continue
        if line.startswith("#question:"):
            finalize_question()
            current_question = line.split(":", 1)[1].strip()
            continue
        if line.startswith("*"):
            option = line[1:].strip()
            if not option:
                raise ValueError("To'g'ri javob bo'sh bo'lishi mumkin emas.")
            current_options.append(option)
            if correct_index is not None:
                raise ValueError(f"'{current_question}' savolida bir nechta to'g'ri javob berilgan.")
            correct_index = len(current_options) - 1
            continue
        if line.startswith("-"):
            option = line[1:].strip()
            if not option:
                raise ValueError("Variant matni bo'sh bo'lishi mumkin emas.")
            current_options.append(option)
            continue
        raise ValueError(f"Noto'g'ri qator formati: {line}")

    finalize_question()

    if not subject:
        raise ValueError("Faylda '#subject: Fan nomi' qatori bo'lishi shart.")
    if not questions:
        raise ValueError("Faylda hech bo'lmaganda bitta savol bo'lishi shart.")

    return subject, short_name, questions


async def pick_next_50_question_ids(
    user_id: int,
    subject: str,
    full_name: Optional[str] = None,
    username: Optional[str] = None,
) -> List[int]:
    all_count = len(subject_tests[subject])
    all_ids = list(range(all_count))
    random.shuffle(all_ids)

    if all_count <= QUESTIONS_PER_SESSION:
        return all_ids

    subject_record = await ensure_user_subject_record_db(user_id, subject, full_name=full_name, username=username)
    used_ids_raw = subject_record.get("used_question_ids", [])
    if isinstance(used_ids_raw, str):
        used_ids_raw = json.loads(used_ids_raw)
    used_ids: Set[int] = set(used_ids_raw)

    remaining_ids = [qid for qid in all_ids if qid not in used_ids]
    random.shuffle(remaining_ids)

    if len(remaining_ids) >= QUESTIONS_PER_SESSION:
        selected = remaining_ids[:QUESTIONS_PER_SESSION]
        used_ids.update(selected)
    else:
        selected = remaining_ids[:]
        unused_after_reset = [qid for qid in all_ids if qid not in set(selected)]
        random.shuffle(unused_after_reset)
        need = QUESTIONS_PER_SESSION - len(selected)
        refill = unused_after_reset[:need]
        selected.extend(refill)
        used_ids = set(selected)

    subject_record["used_question_ids"] = list(used_ids)
    await save_user_subject(user_id, subject, subject_record)
    return selected


async def start_subject_session(
    user_id: int,
    subject: str,
    full_name: Optional[str] = None,
    username: Optional[str] = None,
) -> UserSession:
    if subject not in subject_tests or not subject_tests[subject]:
        raise ValueError(f"{subject} fanida testlar mavjud emas.")

    session = get_or_create_session(user_id)
    selected_ids = await pick_next_50_question_ids(user_id, subject, full_name=full_name, username=username)

    session.subject = subject
    session.question_ids = selected_ids
    session.current_index = 0
    session.score = 0
    session.waiting_poll = False
    session.current_question_id = None
    session.current_poll_id = None
    session.current_poll_message_id = None
    session.current_control_message_id = None
    session.started_at = time.time()
    return session


async def get_subject_rank(subject: str, user_id: int) -> tuple[int, int]:
    rows = await get_all_users_with_subjects()
    leaderboard = []
    for row in rows:
        if row["subject"] != subject:
            continue
        leaderboard.append((
            row["user_id"],
            row["full_name"],
            row["best_percent"],
            row["best_score"],
            row["best_time_seconds"],
            row["best_achieved_at"],
        ))

    leaderboard.sort(key=lambda x: (-x[2], -x[3], x[4], x[5]))
    total_users = len(leaderboard)

    for idx, item in enumerate(leaderboard, start=1):
        if item[0] == user_id:
            return idx, total_users

    return total_users, total_users


async def get_subject_top_10(subject: str, limit: int = 10) -> List[tuple]:
    rows = await get_all_users_with_subjects()
    result = []
    for row in rows:
        if row["subject"] != subject:
            continue
        result.append((
            row["full_name"],
            row["best_percent"],
            row["best_score"],
            row["best_total"],
            row["best_time_seconds"],
            row["best_achieved_at"],
        ))
    result.sort(key=lambda x: (-x[1], -x[2], x[4], x[5]))
    return result[:limit]


async def get_global_leaderboard(limit: int = 10) -> List[tuple]:
    rows = await get_all_users_with_subjects()

    user_bests: Dict[int, dict] = {}
    for row in rows:
        uid = row["user_id"]
        p = row["best_percent"] or 0.0
        s = row["best_score"] or 0
        t = row["best_total"] or 0
        tm = row["best_time_seconds"] or 999999
        ach = row["best_achieved_at"] or 9999999999
        name = row["full_name"]

        if uid not in user_bests:
            user_bests[uid] = {"name": name, "p": p, "s": s, "t": t, "tm": tm, "ach": ach}
        else:
            current = (p, s, -tm, -ach)
            saved = (user_bests[uid]["p"], user_bests[uid]["s"], -user_bests[uid]["tm"], -user_bests[uid]["ach"])
            if current > saved:
                user_bests[uid] = {"name": name, "p": p, "s": s, "t": t, "tm": tm, "ach": ach}

    result = [
        (v["name"], v["p"], v["s"], v["t"], v["tm"], v["ach"])
        for v in user_bests.values()
    ]
    result.sort(key=lambda x: (-x[1], -x[2], x[4], x[5]))
    return result[:limit]


async def save_user_result(
    user_id: int,
    full_name: str,
    subject: str,
    score: int,
    total: int,
    username: Optional[str] = None,
    elapsed_seconds: int = 0,
) -> None:
    percent = round((score / total) * 100, 1) if total else 0.0

    subject_record = await ensure_user_subject_record_db(
        user_id, subject, full_name=full_name, username=username
    )

    subject_record["attempts"] = subject_record.get("attempts", 0) + 1
    subject_record["last_score"] = score
    subject_record["last_total"] = total
    subject_record["last_percent"] = percent
    subject_record["last_time_seconds"] = elapsed_seconds

    old_best_percent = subject_record.get("best_percent", 0.0) or 0.0
    old_best_score = subject_record.get("best_score", 0) or 0
    old_best_time = subject_record.get("best_time_seconds", 0) or 0

    is_better = False
    if percent > old_best_percent:
        is_better = True
    elif percent == old_best_percent and score > old_best_score:
        is_better = True
    elif percent == old_best_percent and score == old_best_score:
        if old_best_time == 0 or elapsed_seconds < old_best_time:
            is_better = True

    if is_better:
        subject_record["best_score"] = score
        subject_record["best_total"] = total
        subject_record["best_percent"] = percent
        subject_record["best_time_seconds"] = elapsed_seconds
        subject_record["best_achieved_at"] = int(time.time())

    await save_user_subject(user_id, subject, subject_record)


async def clear_old_control_message(bot: Bot, user_id: int, session: UserSession):
    if session.current_control_message_id:
        try:
            await bot.delete_message(chat_id=user_id, message_id=session.current_control_message_id)
        except Exception:
            pass
        session.current_control_message_id = None


async def ensure_access(message: Message) -> bool:
    full_name = message.from_user.full_name or message.from_user.first_name or "Foydalanuvchi"
    username = message.from_user.username or ""

    await ensure_user_record_db(message.from_user.id, full_name=full_name, username=username)

    if await is_user_blocked(message.from_user.id):
        await message.answer(
            "⛔ Siz botdan foydalanish uchun bloklangansiz.\nAdmin bilan bog'laning.",
            reply_markup=get_main_reply_keyboard(),
        )
        return False

    # Bloklangan bo'lmasa — admin, tasdiqlangan yoki bepul limitda bo'lsa ruxsat
    user_id = message.from_user.id
    if is_admin(user_id) or await is_user_approved(user_id):
        return True

    # Bepul limit tekshiruvi
    free_used = await get_free_questions_used(user_id)
    if free_used < FREE_QUESTIONS_LIMIT:
        return True

    # Limit tugagan
    payment_info = PAYMENT_INFO.replace("\\n", "\n").replace("\n", "\n")
    await message.answer(
        "⚠️ <b>Bepul savollaringiz tugagan!</b>\n\n"
        "Botdan to'liq foydalanish uchun to'lov qiling va kvitansiyani yuboring:\n\n"
        f"{payment_info}",
        reply_markup=build_payment_keyboard(),
    )
    return False


async def send_welcome(message: Message):
    full_name = message.from_user.full_name or message.from_user.first_name or "Foydalanuvchi"
    username = message.from_user.username or ""

    await ensure_user_record_db(message.from_user.id, full_name=full_name, username=username)

    reply_keyboard = get_main_reply_keyboard()

    if await is_user_blocked(message.from_user.id):
        await message.answer(
            f"Assalomu alaykum, <b>{full_name}</b>.\n\n"
            "⛔ Siz botdan foydalanish uchun bloklangansiz.",
            reply_markup=reply_keyboard,
        )
        return

    subjects_text, keyboard = build_subjects_text_and_keyboard()

    if keyboard is None:
        await message.answer(
            f"Assalomu alaykum, <b>{full_name}</b>.\n\nHozircha admin tomonidan birorta fan yuklanmagan.",
            reply_markup=reply_keyboard,
        )
        return

    # Admin, tasdiqlangan yoki bepul limit tekshiruvi
    user_id = message.from_user.id
    is_adm = is_admin(user_id)
    is_appr = await is_user_approved(user_id)
    free_used = await get_free_questions_used(user_id)
    free_left = max(0, FREE_QUESTIONS_LIMIT - free_used)

    if is_adm or is_appr:
        trial_info = "✅ <b>To'liq foydalanish</b> huquqi berilgan."
        kb_to_show = keyboard
    elif free_left > 0:
        trial_info = f"🎁 Sizda <b>{free_left} ta bepul savol</b> imkoniyati bor.\nUshbu imkoniyat tugagach, botdan to'liq foydalanish uchun to'lov talab qilinadi."
        kb_to_show = keyboard # Bepul savollar bo'lsa fanlarni ko'rsatamiz
    else:
        # Limit tugagan — to'lov xabari
        payment_info = PAYMENT_INFO.replace("\\n", "\n").replace("\n", "\n")
        await message.answer(
            f"Assalomu alaykum, <b>{full_name}</b>.\n\n"
            "⚠️ <b>Bepul savollaringiz tugagan!</b>\n\n"
            "Botdan to'liq foydalanish uchun quyidagi rekvizitlarga to'lov qiling "
            "va kvitansiyani yuboring:\n\n"
            f"{payment_info}",
            reply_markup=build_payment_keyboard(),
        )
        return

    text = (
        f"Assalomu alaykum, <b>{full_name}</b>.\n\n"
        f"{trial_info}\n\n"
        "Quyidagi fanlardan birini tanlang.\n"
        f"Har bir urinishda <b>{QUESTIONS_PER_SESSION}</b> ta savol beriladi.\n"
        "Savollar random tarzda yuboriladi.\n"
        "Test davomida xohlasangiz <b>⏹ Testni tugatish</b> tugmasi bilan yakunlashingiz mumkin."
    )
    await message.answer(text, reply_markup=reply_keyboard)
    if kb_to_show:
        await message.answer(subjects_text if (is_adm or is_appr) else "Fanlar ro'yxati:", reply_markup=kb_to_show)


async def finalize_test(target, user_id: int, full_name: str, username: Optional[str] = None):
    session = get_or_create_session(user_id)

    if not session.subject or not session.question_ids:
        text = "Faol test mavjud emas."
        subjects_text, keyboard = build_subjects_text_and_keyboard()
        reply_keyboard = get_main_reply_keyboard()

        if isinstance(target, Message):
            await target.answer(text, reply_markup=reply_keyboard)
            if keyboard:
                await target.answer(subjects_text, reply_markup=keyboard)
        else:
            await target.send_message(user_id, text, reply_markup=reply_keyboard)
            if keyboard:
                await target.send_message(user_id, subjects_text, reply_markup=keyboard)
        return

    total = len(session.question_ids)
    percent = round((session.score / total) * 100, 1) if total else 0.0
    elapsed_seconds = int(time.time() - session.started_at) if session.started_at else 0

    await save_user_result(
        user_id, full_name, session.subject, session.score, total,
        username=username, elapsed_seconds=elapsed_seconds,
    )
    rank, total_users = await get_subject_rank(session.subject, user_id)
    time_text = format_duration(elapsed_seconds)

    text = (
        f"✅ <b>Test yakunlandi</b>\n\n"
        f"📚 Fan: <b>{session.subject}</b>\n"
        f"👤 Talaba: <b>{full_name}</b>\n"
        f"🎯 Natija: <b>{session.score}/{total}</b>\n"
        f"📈 Foiz: <b>{percent}%</b>\n"
        f"⏱ Ishlagan vaqtingiz: <b>{time_text}</b>\n"
        f"🏆 Reytingdagi o'rningiz: <b>{rank}</b>-o'rin / {total_users} ta foydalanuvchi\n\n"
        "Yana ishlash uchun pastdan fan tanlang."
    )

    subjects_text, keyboard = build_subjects_text_and_keyboard()
    reply_keyboard = get_main_reply_keyboard()

    if isinstance(target, Message):
        await target.answer(text, reply_markup=reply_keyboard)
        if keyboard:
            await target.answer(subjects_text, reply_markup=keyboard)
    else:
        await target.send_message(user_id, text, reply_markup=reply_keyboard)
        if keyboard:
            await target.send_message(user_id, subjects_text, reply_markup=keyboard)

    session.waiting_poll = False
    session.current_question_id = None
    session.current_poll_id = None
    session.current_poll_message_id = None
    session.current_control_message_id = None
    session.subject = None
    session.question_ids = []
    session.current_index = 0
    session.score = 0
    session.started_at = None


async def send_next_question(target, user_id: int):
    session = get_or_create_session(user_id)

    if not session.subject or not session.question_ids:
        text = "Avval fanlardan birini tanlang."
        subjects_text, keyboard = build_subjects_text_and_keyboard()
        reply_keyboard = get_main_reply_keyboard()

        if isinstance(target, Message):
            await target.answer(text, reply_markup=reply_keyboard)
            if keyboard:
                await target.answer(subjects_text, reply_markup=keyboard)
        else:
            await target.send_message(user_id, text, reply_markup=reply_keyboard)
            if keyboard:
                await target.send_message(user_id, subjects_text, reply_markup=keyboard)
        return

    if session.current_index >= len(session.question_ids):
        user_record = await get_user_record(user_id)
        full_name = user_record.get("full_name", "Foydalanuvchi") if user_record else "Foydalanuvchi"
        username = user_record.get("username", "") if user_record else ""
        await finalize_test(target, user_id, full_name, username=username)
        return

    # Bepul limit tekshiruvi — faqat tasdiqlanmagan foydalanuvchilar uchun
    if not await is_user_approved(user_id):
        if await has_free_limit_reached(user_id):
            # Sessiyani to'xtatish
            session.subject = None
            session.question_ids = []
            await clear_old_control_message(
                target if isinstance(target, Bot) else target.bot,
                user_id, session
            )
            payment_info = PAYMENT_INFO.replace("\\n", "\n").replace("\n", "\n")
            text = (
                "🎉 <b>5 ta bepul savolni ishladingiz!</b>\n\n"
                "Botdan to'liq foydalanish uchun quyidagi rekvizitlarga to'lov qiling "
                "va kvitansiyani yuboring — admin tasdiqlagach barcha fanlar ochiladi.\n\n"
                f"{payment_info}"
            )
            kb = build_payment_keyboard() # Kvitansiya va Admin tugmalari
            if isinstance(target, Message):
                await target.answer(text, reply_markup=kb)
            else:
                await target.send_message(user_id, text, reply_markup=kb)
            return
        else:
            # Bepul savolni +1 qilamiz
            await increment_free_questions(user_id)

    question_id = session.question_ids[session.current_index]
    question = subject_tests[session.subject][question_id]

    safe_question = question.question[:300]
    safe_options = [opt[:100] for opt in question.options]
    safe_explanation = f"To'g'ri javob: {safe_options[question.correct_option_id]}"[:200]

    poll_kwargs = dict(
        question=safe_question,
        options=safe_options,
        type=PollType.QUIZ,
        correct_option_id=question.correct_option_id,
        explanation=safe_explanation,
        is_anonymous=False,
        open_period=POLL_OPEN_PERIOD,
    )

    if isinstance(target, Message):
        sent = await target.answer_poll(**poll_kwargs)
        bot = target.bot
    else:
        sent = await target.send_poll(chat_id=user_id, **poll_kwargs)
        bot = target

    session.waiting_poll = True
    session.current_question_id = question_id
    session.current_poll_id = sent.poll.id
    session.current_poll_message_id = sent.message_id
    session.poll_sent_at = time.time()
    poll_to_user[sent.poll.id] = user_id

    await clear_old_control_message(bot, user_id, session)

    control_text = (
        f"📘 <b>{session.subject}</b>\n"
        f"❓ Savol: <b>{session.current_index + 1}/{len(session.question_ids)}</b>\n"
        f"🎯 Hozirgi ball: <b>{session.score}</b>\n\n"
        "Testni yakunlamoqchi bo'lsangiz, tugmani bosing:"
    )
    control_message = await bot.send_message(
        chat_id=user_id,
        text=control_text,
        reply_markup=build_finish_test_keyboard(),
    )
    session.current_control_message_id = control_message.message_id

    asyncio.create_task(poll_timeout_watcher(bot, user_id, sent.poll.id, session.poll_sent_at))


async def poll_timeout_watcher(bot: Bot, user_id: int, poll_id: str, sent_at: float):
    await asyncio.sleep(POLL_OPEN_PERIOD + 5)

    session = user_sessions.get(user_id)
    if not session:
        return
    if session.current_poll_id != poll_id:
        return
    if session.poll_sent_at != sent_at:
        return
    if not session.waiting_poll:
        return

    session.waiting_poll = False
    session.paused = True
    session.current_poll_message_id = None
    session.poll_sent_at = None
    poll_to_user.pop(poll_id, None)

    await clear_old_control_message(bot, user_id, session)

    pause_text = (
        f"⏸ <b>Test pauza qilindi</b>\n\n"
        f"📘 Fan: <b>{session.subject}</b>\n"
        f"❓ Savol: <b>{session.current_index + 1}/{len(session.question_ids)}</b>\n"
        f"🎯 Hozirgi ball: <b>{session.score}</b>\n\n"
        "Vaqt tugadi. Testni davom ettirasizmi yoki tugatmoqchimisiz?"
    )
    control_msg = await bot.send_message(
        chat_id=user_id,
        text=pause_text,
        reply_markup=build_paused_keyboard(),
    )
    session.current_control_message_id = control_msg.message_id


async def set_bot_commands(bot: Bot):
    # Oddiy foydalanuvchilar uchun (hamma ko'radi)
    user_commands = [
        BotCommand(command="start", description="Botni ishga tushirish"),
        BotCommand(command="fans", description="Fanlar ro'yxati"),
        BotCommand(command="reyting", description="Umumiy reyting"),
        BotCommand(command="my_results", description="Mening natijalarim"),
        BotCommand(command="subject_top", description="Fan bo'yicha TOP-10"),
    ]
    await bot.set_my_commands(user_commands, scope=BotCommandScopeAllPrivateChats())

    # Admin uchun qo'shimcha buyruqlar (faqat admin chatida ko'rinadi)
    admin_commands = user_commands + [
        BotCommand(command="admin", description="Admin panel"),
        BotCommand(command="template", description="Test shabloni"),
        BotCommand(command="approve", description="Foydalanuvchiga ruxsat berish"),
        BotCommand(command="block", description="Foydalanuvchini bloklash"),
        BotCommand(command="unblock", description="Blokdan chiqarish"),
        BotCommand(command="pending_users", description="Tasdiqlanmagan userlar"),
        BotCommand(command="approved_users", description="Tasdiqlangan userlar ro'yxati"),
        BotCommand(command="remove_rating", description="Reytingdan o'chirish"),
        BotCommand(command="delete_user", description="Foydalanuvchini to'liq o'chirish"),
        BotCommand(command="delete_subject", description="Fanni o'chirish"),
        BotCommand(command="rename_subject", description="Fan nomini o'zgartirish"),
        BotCommand(command="approve_template", description="Ruxsat berish shabloni (.xlsx)"),
        BotCommand(command="broadcast", description="Barcha foydalanuvchilarga xabar"),
    ]
    for admin_id in ADMIN_IDS:
        try:
            await bot.set_my_commands(
                admin_commands,
                scope=BotCommandScopeChat(chat_id=admin_id)
            )
        except Exception:
            pass


# ============================================================
# 8) FOYDALANUVCHI BUYRUQLARI
# ============================================================
@router.message(CommandStart())
async def cmd_start(message: Message):
    await send_welcome(message)


@router.message(Command("fans"))
async def cmd_fans(message: Message):
    if not await ensure_access(message):
        return
    subjects_text, keyboard = build_subjects_text_and_keyboard()
    if not keyboard:
        await message.answer("Hozircha birorta fan yuklanmagan.", reply_markup=get_main_reply_keyboard())
        return
    await message.answer(subjects_text, reply_markup=get_main_reply_keyboard())
    await message.answer("Fanni tanlang:", reply_markup=keyboard)


@router.message(Command("reyting"))
async def cmd_reyting(message: Message):
    if not await ensure_access(message):
        return
    top = await get_global_leaderboard(limit=10)
    if not top:
        await message.answer("Hali reyting shakllanmagan.", reply_markup=get_main_reply_keyboard())
        return
    lines = ["🏆 <b>Umumiy reyting TOP-10</b>\n"]
    for idx, (name, percent, score, total, time_sec, _) in enumerate(top, start=1):
        lines.append(
            f"{idx}. <b>{name}</b> — {score}/{total} ({percent}%) — ⏱ {format_duration(time_sec)}"
        )
    await message.answer("\n".join(lines), reply_markup=get_main_reply_keyboard())


@router.message(Command("my_results"))
async def cmd_my_results(message: Message):
    if not await ensure_access(message):
        return
    user_record = await get_user_record(message.from_user.id)
    subject_rows = await get_user_all_subjects(message.from_user.id)
    if not user_record or not subject_rows:
        await message.answer("Sizda hali saqlangan natijalar yo'q.", reply_markup=get_main_reply_keyboard())
        return
    lines = [f"📊 <b>{user_record.get('full_name', 'Foydalanuvchi')}</b> natijalari:\n"]
    for row in sorted(subject_rows, key=lambda x: x["subject"]):
        lines.append(
            f"📘 <b>{row['subject']}</b>\n"
            f"  • Urinishlar: {row.get('attempts', 0)}\n"
            f"  • Oxirgi natija: {row.get('last_score', 0)}/{row.get('last_total', 0)} "
            f"({row.get('last_percent', 0)}%) — ⏱ {format_duration(row.get('last_time_seconds', 0))}\n"
            f"  • Eng yaxshi natija: {row.get('best_score', 0)}/{row.get('best_total', 0)} "
            f"({row.get('best_percent', 0)}%) — ⏱ {format_duration(row.get('best_time_seconds', 0))}\n"
        )
    await message.answer("\n".join(lines), reply_markup=get_main_reply_keyboard())


@router.message(Command("subject_top"))
async def cmd_subject_top(message: Message):
    if not await ensure_access(message):
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Foydalanish: /subject_top Fan nomi", reply_markup=get_main_reply_keyboard())
        return
    subject = parts[1].strip()
    if subject not in subject_tests:
        await message.answer("Bunday fan topilmadi.", reply_markup=get_main_reply_keyboard())
        return
    top = await get_subject_top_10(subject, limit=10)
    if not top:
        await message.answer(f"{subject} fanidan hali reyting yo'q.", reply_markup=get_main_reply_keyboard())
        return
    lines = [f"🏆 <b>{subject}</b> fanidan TOP-10\n"]
    lines.append("Saralash mezoni: foiz → to'g'ri javob → vaqt → birinchi erishgan vaqt\n")
    for idx, (name, percent, score, total, time_sec, _) in enumerate(top, start=1):
        lines.append(
            f"{idx}. <b>{name}</b> — {score}/{total} ({percent}%) — ⏱ {format_duration(time_sec)}"
        )
    await message.answer("\n".join(lines), reply_markup=get_main_reply_keyboard())


@router.message(F.text == "📚 Fanlar")
async def btn_fans(message: Message):
    await cmd_fans(message)


@router.message(F.text == "🏆 Reyting")
async def btn_reyting(message: Message):
    await cmd_reyting(message)


@router.message(F.text == "📊 Mening natijam")
async def btn_my_results(message: Message):
    await cmd_my_results(message)


@router.message(F.text == "ℹ️ Yordam")
async def btn_help(message: Message):
    full_name = message.from_user.full_name or message.from_user.first_name or "Foydalanuvchi"
    username = message.from_user.username or ""
    await ensure_user_record_db(message.from_user.id, full_name=full_name, username=username)

    if await is_user_blocked(message.from_user.id):
        await message.answer(
            "⛔ Siz botdan foydalanish uchun bloklangansiz.",
            reply_markup=get_main_reply_keyboard(),
        )
        return

    if not await is_user_approved(message.from_user.id):
        await message.answer(
            "🔒 Botdan foydalanish uchun to'lov talab etiladi.\nAdmin bilan bog'laning.",
            reply_markup=build_access_request_keyboard(),
        )
        return

    text = (
        "ℹ️ <b>Yordam</b>\n\n"
        "1. <b>📚 Fanlar</b> ni bosing\n"
        "2. Fanni tanlang\n"
        f"3. Sizga {QUESTIONS_PER_SESSION} ta savol yuboriladi\n"
        "4. Xohlasangiz test paytida <b>⏹ Testni tugatish</b> tugmasi bilan yakunlashingiz mumkin\n"
        "5. Test oxirida natija, vaqt va reytingdagi o'rningiz ko'rsatiladi\n\n"
        "Qo'shimcha buyruqlar:\n"
        "/start\n"
        "/fans\n"
        "/reyting\n"
        "/my_results\n"
        "/subject_top Fan nomi"
    )
    await message.answer(text, reply_markup=get_main_reply_keyboard())


# ============================================================
# 9) CALLBACKLAR
# ============================================================
@router.callback_query(F.data == "request_access")
async def request_access_handler(callback: CallbackQuery):
    user = callback.from_user
    full_name = user.full_name or user.first_name or "Foydalanuvchi"
    username = user.username or ""

    user_record = await ensure_user_record_db(user.id, full_name=full_name, username=username)

    if user_record.get("is_blocked"):
        await callback.answer("Siz bloklangansiz.", show_alert=True)
        return

    if user_record.get("is_approved"):
        await callback.answer("Siz allaqachon tasdiqlangansiz.", show_alert=True)
        return

    admin_text = (
        "📥 <b>Yangi ruxsat so'rovi</b>\n\n"
        f"👤 Ism: <b>{full_name}</b>\n"
        f"🆔 ID: <code>{user.id}</code>\n"
        "🔗 Username: " + (('@' + username) if username else "yo'q")
    )

    sent_count = 0
    for admin_id in ADMIN_IDS:
        try:
            await callback.bot.send_message(
                admin_id,
                admin_text,
                reply_markup=build_admin_approval_keyboard(user.id),
            )
            sent_count += 1
        except Exception:
            pass

    if sent_count:
        await callback.answer("So'rovingiz adminga yuborildi.", show_alert=True)
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except Exception:
            pass
        await callback.message.answer(
            "✅ So'rovingiz yuborildi.\nAdmin tasdiqlagandan keyin botdan foydalanishingiz mumkin."
        )
    else:
        await callback.answer("Adminga yuborib bo'lmadi.", show_alert=True)


@router.callback_query(F.data.startswith("approve_user:"))
async def approve_user_handler(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("Bu amal faqat admin uchun.", show_alert=True)
        return

    user_id = int(callback.data.split(":", 1)[1])
    await ensure_user_record_db(user_id)
    await set_user_approved(user_id, approved=True, blocked=False)

    try:
        await callback.bot.send_message(
            user_id,
            "✅ Siz admin tomonidan tasdiqlandingiz.\nEndi botdan foydalanishingiz mumkin.\n/start"
        )
    except Exception:
        pass

    await callback.answer("Foydalanuvchiga ruxsat berildi.")
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass
    await callback.message.answer(f"✅ User <code>{user_id}</code> tasdiqlandi.")


@router.callback_query(F.data.startswith("reject_user:"))
async def reject_user_handler(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("Bu amal faqat admin uchun.", show_alert=True)
        return

    user_id = int(callback.data.split(":", 1)[1])
    await ensure_user_record_db(user_id)
    await set_user_approved(user_id, approved=False)

    try:
        await callback.bot.send_message(
            user_id,
            "❌ Admin sizning botdan foydalanish so'rovingizni rad etdi."
        )
    except Exception:
        pass

    await callback.answer("So'rov rad etildi.")
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass
    await callback.message.answer(f"❌ User <code>{user_id}</code> so'rovi rad etildi.")


@router.callback_query(F.data.startswith("block_user:"))
async def block_user_handler(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer("Bu amal faqat admin uchun.", show_alert=True)
        return

    user_id = int(callback.data.split(":", 1)[1])
    await ensure_user_record_db(user_id)
    await set_user_blocked(user_id, blocked=True)

    try:
        await callback.bot.send_message(
            user_id,
            "⛔ Siz admin tomonidan bloklandingiz.\nBotdan foydalana olmaysiz."
        )
    except Exception:
        pass

    await callback.answer("Foydalanuvchi bloklandi.")
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass
    await callback.message.answer(f"⛔ User <code>{user_id}</code> bloklandi.")


@router.callback_query(F.data == "refresh_subjects")
async def refresh_subjects_handler(callback: CallbackQuery):
    subjects_text, keyboard = build_subjects_text_and_keyboard()
    if keyboard is None:
        await callback.answer("Hozircha fanlar yuklanmagan.", show_alert=True)
        return
    try:
        await callback.message.edit_text(subjects_text, reply_markup=keyboard)
    except Exception:
        pass
    await callback.answer("Fanlar yangilandi.")


@router.callback_query(F.data == "continue_test")
async def continue_test_handler(callback: CallbackQuery):
    user_id = callback.from_user.id
    session = get_or_create_session(user_id)

    if not session.subject or not session.question_ids:
        await callback.answer("Faol test topilmadi.", show_alert=True)
        return

    if not session.paused:
        await callback.answer("Test allaqachon davom etmoqda.", show_alert=True)
        return

    session.paused = False
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass

    await callback.answer("Test davom ettirildi!")
    await send_next_question(callback.message, user_id)



@router.callback_query(F.data == "send_receipt")
async def send_receipt_handler(callback: CallbackQuery):
    user_id = callback.from_user.id
    full_name = callback.from_user.full_name or callback.from_user.first_name or "Foydalanuvchi"

    await callback.answer()
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass

    await callback.message.answer(
        "📎 <b>Kvitansiyani yuboring</b>\n\n"
        "To'lov kvitansiyasini <b>rasm</b> yoki <b>fayl</b> sifatida yuboring.\n"
        "Admin ko'rib chiqib tasdiqleydi."
    )


@router.message(F.photo | F.document & F.document.mime_type.startswith("image"))
async def receipt_photo_handler(message: Message, bot: Bot):
    user_id = message.from_user.id
    full_name = message.from_user.full_name or message.from_user.first_name or "Foydalanuvchi"
    username = message.from_user.username or ""

    # Admin bo'lsa o'tkazib yuborish
    if is_admin(user_id):
        return

    # Foydalanuvchi allaqachon tasdiqlangan bo'lsa o'tkazib yuborish
    if await is_user_approved(user_id):
        return

    # Adminga kvitansiyani yuborish
    caption = (
        f"💳 <b>Yangi to'lov kvitansiyasi</b>\n\n"
        f"👤 Ism: <b>{full_name}</b>\n"
        f"🆔 ID: <code>{user_id}</code>\n"
        f"🔗 Username: {('@' + username) if username else 'yo\'q'}"
    )

    sent_count = 0
    for admin_id in ADMIN_IDS:
        try:
            if message.photo:
                await bot.send_photo(
                    admin_id,
                    photo=message.photo[-1].file_id,
                    caption=caption,
                    reply_markup=build_admin_approval_keyboard(user_id),
                )
            else:
                await bot.send_document(
                    admin_id,
                    document=message.document.file_id,
                    caption=caption,
                    reply_markup=build_admin_approval_keyboard(user_id),
                )
            sent_count += 1
        except Exception:
            pass

    if sent_count:
        await message.answer(
            "✅ Kvitansiyangiz adminga yuborildi!\n"
            "Admin ko'rib chiqib tasdiqlagach xabar olasiz."
        )
    else:
        await message.answer("❌ Adminga yuborib bo'lmadi. Keyinroq urinib ko'ring.")

@router.callback_query(F.data == "finish_test")
async def finish_test_handler(callback: CallbackQuery):
    user_id = callback.from_user.id
    session = get_or_create_session(user_id)

    if not session.subject or not session.question_ids:
        await callback.answer("Faol test topilmadi.", show_alert=True)
        return

    if session.current_poll_message_id:
        try:
            await callback.bot.stop_poll(chat_id=user_id, message_id=session.current_poll_message_id)
        except Exception:
            pass

    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass

    full_name = callback.from_user.full_name or callback.from_user.first_name or "Foydalanuvchi"
    username = callback.from_user.username or ""
    await callback.answer("Test yakunlandi.")
    await finalize_test(callback.message, user_id, full_name, username=username)


@router.callback_query(F.data.startswith("subj:"))
async def subject_selected_callback(callback: CallbackQuery):
    full_name = callback.from_user.full_name or callback.from_user.first_name or "Foydalanuvchi"
    username = callback.from_user.username or ""

    await ensure_user_record_db(callback.from_user.id, full_name=full_name, username=username)

    if await is_user_blocked(callback.from_user.id):
        await callback.answer("Siz bloklangansiz.", show_alert=True)
        return

    uid = callback.from_user.id
    if not is_admin(uid) and not await is_user_approved(uid):
        free_used = await get_free_questions_used(uid)
        if free_used >= FREE_QUESTIONS_LIMIT:
            payment_info = PAYMENT_INFO.replace("\\n", "\n").replace("\n", "\n")
            await callback.answer("Bepul savollar tugagan!", show_alert=True)
            await callback.message.answer(
                "⚠️ <b>Bepul savollaringiz tugagan!</b>\n\n"
                f"{payment_info}",
                reply_markup=build_payment_keyboard(),
            )
            return

    try:
        idx = int(callback.data.split("subj:", 1)[1])
        subject = sorted(subject_tests.keys())[idx]
    except (ValueError, IndexError):
        await callback.answer("Bu fan topilmadi yoki o'chirib yuborilgan.", show_alert=True)
        return

    if subject not in subject_tests:
        await callback.answer("Bu fan topilmadi yoki o'chirib yuborilgan.", show_alert=True)
        return

    session = await start_subject_session(callback.from_user.id, subject, full_name=full_name, username=username)

    await callback.message.answer(
        f"📘 <b>{subject}</b> fani tanlandi.\n"
        f"Jami <b>{len(session.question_ids)}</b> ta savol yuboriladi.\n"
        f"👤 Talaba: <b>{full_name}</b>"
    )
    await callback.answer()
    await send_next_question(callback.message, callback.from_user.id)


# ============================================================
# 10) POLL JAVOBLARI
# ============================================================
@router.poll_answer()
async def handle_poll_answer(poll_answer: PollAnswer, bot: Bot):
    user_id = poll_to_user.get(poll_answer.poll_id)
    if not user_id:
        return

    session = user_sessions.get(user_id)
    if not session or not session.subject or session.current_question_id is None:
        return
    if session.current_poll_id != poll_answer.poll_id:
        return

    question = subject_tests[session.subject][session.current_question_id]
    selected = poll_answer.option_ids

    if selected and selected[0] == question.correct_option_id:
        session.score += 1

    old_message_id = session.current_poll_message_id

    session.current_index += 1
    session.waiting_poll = False
    session.current_question_id = None
    session.current_poll_id = None
    session.current_poll_message_id = None
    session.poll_sent_at = None

    if old_message_id is not None:
        try:
            await bot.stop_poll(chat_id=user_id, message_id=old_message_id)
        except Exception:
            pass

    poll_to_user.pop(poll_answer.poll_id, None)
    await clear_old_control_message(bot, user_id, session)
    await asyncio.sleep(0.3)
    await send_next_question(bot, user_id)


# ============================================================
# 11) ADMIN BUYRUQLARI
# ============================================================
@router.message(Command("template"))
async def cmd_template(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    file = BufferedInputFile(TEMPLATE_TEXT.encode("utf-8"), filename="test_template.txt")
    await message.answer_document(
        document=file,
        caption=(
            "Shu shablon bo'yicha test tayyorlang va botga txt fayl sifatida yuboring.\n\n"
            "<b>#subject:</b> — fanning to'liq nomi\n"
            "<b>#short:</b> — tugmada chiqadigan qisqa nom (ixtiyoriy)\n\n"
            "Bot fan nomini fayldagi #subject dan oladi va testlarni o'zida saqlaydi."
        ),
    )


@router.message(Command("delete_subject"))
async def cmd_delete_subject(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Foydalanish: /delete_subject Fan nomi")
        return
    subject = parts[1].strip()
    if subject not in subject_tests:
        await message.answer("Bunday fan topilmadi.")
        return
    del subject_tests[subject]
    subject_short_names.pop(subject, None)
    await delete_subject_db(subject)
    await message.answer(f"🗑 <b>{subject}</b> fani o'chirildi.")


@router.message(Command("rename_subject"))
async def cmd_rename_subject(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2 or "|" not in parts[1]:
        await message.answer(
            "Foydalanish:\n"
            "/rename_subject Eski nom | Yangi to'liq nom | Yangi qisqa nom\n\n"
            "Faqat to'liq nomni o'zgartirish:\n"
            "/rename_subject Eski nom | Yangi to'liq nom\n\n"
            "Faqat qisqa nomni o'zgartirish:\n"
            "/rename_subject Eski nom | | Yangi qisqa nom"
        )
        return

    raw = parts[1].split("|")
    old_name = raw[0].strip()
    new_full = raw[1].strip() if len(raw) > 1 else ""
    new_short = raw[2].strip() if len(raw) > 2 else None

    if not old_name:
        await message.answer("Eski nom bo'sh bo'lishi mumkin emas.")
        return
    if old_name not in subject_tests:
        await message.answer(f"❌ <b>{old_name}</b> nomli fan topilmadi.")
        return
    if new_full and new_full != old_name and new_full in subject_tests:
        await message.answer(f"❌ <b>{new_full}</b> nomli fan allaqachon mavjud.")
        return

    questions = subject_tests[old_name]
    current_short = subject_short_names.get(old_name, "")

    final_full = new_full if new_full else old_name
    final_short = new_short if new_short is not None else current_short

    pool = await get_pool()
    async with pool.acquire() as conn:
        questions_json = json.dumps([q.to_dict() for q in questions], ensure_ascii=False)
        await conn.execute("""
            INSERT INTO subjects (subject, short_name, questions)
            VALUES ($1, $2, $3::jsonb)
            ON CONFLICT (subject) DO UPDATE SET
                short_name = EXCLUDED.short_name,
                questions = EXCLUDED.questions
        """, final_full, final_short, questions_json)
        await conn.execute(
            "UPDATE user_subjects SET subject = $1 WHERE subject = $2",
            final_full, old_name
        )
        if final_full != old_name:
            await conn.execute("DELETE FROM subjects WHERE subject = $1", old_name)

    if final_full != old_name:
        del subject_tests[old_name]
        subject_short_names.pop(old_name, None)

    subject_tests[final_full] = questions
    subject_short_names[final_full] = final_short

    lines = ["✅ Fan muvaffaqiyatli yangilandi:\n"]
    if new_full and new_full != old_name:
        lines.append(f"📝 To'liq nom: <b>{old_name}</b> → <b>{final_full}</b>")
    if new_short is not None:
        lines.append(f"🔤 Qisqa nom: <b>{current_short or '(yo\'q)'}</b> → <b>{final_short or '(yo\'q)'}</b>")
    await message.answer("\n".join(lines))


@router.message(Command("approve"))
async def cmd_approve(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Foydalanish: /approve USER_ID")
        return
    user_id = int(parts[1])
    await ensure_user_record_db(user_id)
    await set_user_approved(user_id, approved=True, blocked=False)
    try:
        await message.bot.send_message(
            user_id,
            "✅ Siz admin tomonidan tasdiqlandingiz.\nEndi botdan foydalanishingiz mumkin.\n/start"
        )
    except Exception:
        pass
    await message.answer(f"✅ {user_id} foydalanuvchiga ruxsat berildi.")


@router.message(Command("unblock"))
async def cmd_unblock(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Foydalanish: /unblock USER_ID")
        return
    user_id = int(parts[1])
    await ensure_user_record_db(user_id)
    await set_user_blocked(user_id, blocked=False)
    await message.answer(f"✅ {user_id} foydalanuvchi blokdan chiqarildi.")


@router.message(Command("block"))
async def cmd_block(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Foydalanish: /block USER_ID")
        return
    user_id = int(parts[1])
    await ensure_user_record_db(user_id)
    await set_user_blocked(user_id, blocked=True)
    try:
        await message.bot.send_message(
            user_id,
            "⛔ Siz admin tomonidan bloklandingiz.\nBotdan foydalana olmaysiz."
        )
    except Exception:
        pass
    await message.answer(f"⛔ {user_id} foydalanuvchi bloklandi.")


@router.message(Command("pending_users"))
async def cmd_pending_users(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    pending = await get_pending_users()
    if not pending:
        await message.answer("Tasdiqlanmagan userlar yo'q.")
        return
    for udata in pending:
        uid = udata["user_id"]
        text = (
            "🕓 <b>Tasdiqlanmagan foydalanuvchi</b>\n\n"
            f"👤 Ism: <b>{udata.get('full_name', 'Foydalanuvchi')}</b>\n"
            f"🆔 ID: <code>{uid}</code>\n"
            "🔗 Username: " + (('@' + udata.get('username')) if udata.get('username') else "yo'q")
        )
        await message.answer(text, reply_markup=build_admin_approval_keyboard(uid))


@router.message(Command("approved_users"))
async def cmd_approved_users(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    users = await get_all_approved_users()
    if not users:
        await message.answer("Hozircha tasdiqlangan foydalanuvchi yo'q.")
        return
    lines = [f"✅ <b>Tasdiqlangan foydalanuvchilar ({len(users)} ta):</b>\n"]
    for i, u in enumerate(users, start=1):
        username = f"@{u['username']}" if u.get("username") else "yo'q"
        lines.append(
            f"{i}. <b>{u.get('full_name', 'Foydalanuvchi')}</b>\n"
            f"   🆔 <code>{u['user_id']}</code> | 🔗 {username}"
        )
    await message.answer("\n".join(lines))


@router.message(Command("remove_rating"))
async def cmd_remove_rating(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Foydalanish: /remove_rating USER_ID")
        return
    user_id = int(parts[1])
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("DELETE FROM user_subjects WHERE user_id = $1", user_id)
    await message.answer(f"🗑 <code>{user_id}</code> foydalanuvchi reytingdan o'chirildi.")


@router.message(Command("delete_user"))
async def cmd_delete_user(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Foydalanish: /delete_user USER_ID")
        return
    user_id = int(parts[1])
    user = await get_user_record(user_id)
    if not user:
        await message.answer(f"❌ <code>{user_id}</code> foydalanuvchi topilmadi.")
        return
    await delete_user_db(user_id)
    try:
        await message.bot.send_message(
            user_id,
            "🗑 Sizning akkountingiz o'chirildi.\n"
            "Qayta foydalanish uchun /start bosing va ruxsat so'rang."
        )
    except Exception:
        pass
    await message.answer(f"🗑 <code>{user_id}</code> foydalanuvchi to'liq o'chirildi.")



@router.message(Command("reset_trial"))
async def cmd_reset_trial(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].isdigit():
        await message.answer("Foydalanish: /reset_trial USER_ID")
        return
    user_id = int(parts[1])
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE users SET free_questions_used = 0 WHERE user_id = $1", user_id
        )
    await message.answer(f"✅ <code>{user_id}</code> foydalanuvchining bepul savollari tiklandi.")

@router.message(Command("admin"))
async def cmd_admin(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    text = (
        "<b>⚙️ Admin buyruqlari:</b>\n\n"
        "📚 <b>Fanlar boshqaruvi:</b>\n"
        "/fans — barcha fanlar ro'yxatini ko'rish\n"
        "/template — test yuklash uchun .txt shablon olish\n"
        "/delete_subject <i>Fan nomi</i> — fanni o'chirish\n"
        "  Misol: <code>/delete_subject Matematika</code>\n"
        "/rename_subject <i>Eski | Yangi toliq | Yangi qisqa</i> — fan nomini ozgartirish\n"
        "  Misol: <code>/rename_subject Matematika | Oliy matematika | OliyMat</code>\n\n"
        "👥 <b>Foydalanuvchilar boshqaruvi:</b>\n"
        "/pending_users — ruxsat kutayotgan userlar royxati\n"
        "/approved_users — tasdiqlangan foydalanuvchilar royxati\n"
        "/approve <i>USER_ID</i> — foydalanuvchiga ruxsat berish\n"
        "  Misol: <code>/approve 123456789</code>\n"
        "/approve_template — ommaviy ruxsat berish uchun .xlsx shablon olish\n"
        "  (Faylni toldirib botga yuboring — hammaga avtomatik ruxsat beriladi)\n"
        "/block <i>USER_ID</i> — foydalanuvchini bloklash\n"
        "  Misol: <code>/block 123456789</code>\n"
        "/unblock <i>USER_ID</i> — foydalanuvchini blokdan chiqarish\n"
        "  Misol: <code>/unblock 123456789</code>\n"
        "/delete_user <i>USER_ID</i> — foydalanuvchini toliq ochirish\n"
        "  (barcha natijalari ham ochadi, qayta ruxsat sorashi kerak boladi)\n"
        "  Misol: <code>/delete_user 123456789</code>\n"
        "/remove_rating <i>USER_ID</i> — faqat reytingini ochirish (akkaunt qoladi)\n"
        "  Misol: <code>/remove_rating 123456789</code>\n\n"
        "📊 <b>Reyting:</b>\n"
        "/reyting — umumiy reyting\n"
        "/subject_top <i>Fan nomi</i> — fan boyicha TOP-10\n"
        "  Misol: <code>/subject_top Matematika</code>\n\n"
        "📢 <b>Xabar yuborish:</b>\n"
        "/broadcast — barcha foydalanuvchilarga xabar yuborish\n"
        "  (buyruqdan keyin istalgan matn, rasm yoki fayl yuboring)\n\n"
        "Yangi foydalanuvchilar test ishlashdan oldin admin ruxsati olishi kerak."
    )
    await message.answer(text)


@router.message(Command("approve_template"))
async def cmd_approve_template(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    ws.column_dimensions["A"].width = 20
    ws["A1"] = "user_id"
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    for i, sample_id in enumerate([123456789, 987654321, 111222333], start=2):
        ws[f"A{i}"] = sample_id

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    file = BufferedInputFile(buf.read(), filename="approve_template.xlsx")
    await message.answer_document(
        document=file,
        caption=(
            "📋 <b>Ruxsat berish shabloni</b>\n\n"
            "1. Faylni oching\n"
            "2. <b>A</b> ustuniga foydalanuvchilarning Telegram ID larini yozing\n"
            "3. Birinchi qator (<b>user_id</b>) o'zgarmаsin\n"
            "4. Faylni botga yuboring — hammaga avtomatik ruxsat beriladi"
        ),
    )


async def process_approve_xlsx(bot: Bot, file_bytes: bytes, message: Message):
    buf = BytesIO(file_bytes)
    try:
        wb = openpyxl.load_workbook(buf, read_only=True)
    except Exception:
        await message.answer("❌ Excel faylni o'qib bo'lmadi. To'g'ri .xlsx fayl yuboring.")
        return

    ws = wb.active
    user_ids = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[0] if row else None
        if val is None:
            continue
        try:
            uid = int(str(val).strip())
            user_ids.append(uid)
        except (ValueError, TypeError):
            continue

    if not user_ids:
        await message.answer("❌ Faylda birorta ham to'g'ri user_id topilmadi.")
        return

    await message.answer(f"⏳ {len(user_ids)} ta foydalanuvchiga ruxsat berilmoqda...")

    success = []
    failed = []

    for uid in user_ids:
        try:
            await ensure_user_record_db(uid)
            await set_user_approved(uid, approved=True, blocked=False)
            try:
                await bot.send_message(
                    uid,
                    "✅ Siz admin tomonidan tasdiqlandingiz.\nEndi botdan foydalanishingiz mumkin.\n/start"
                )
            except Exception:
                pass
            success.append(str(uid))
        except Exception:
            failed.append(str(uid))

    lines = []
    if success:
        lines.append(f"✅ Tasdiqlandi — <b>{len(success)}</b> ta foydalanuvchi")
    if failed:
        lines.append(f"❌ Xato — <b>{len(failed)}</b> ta:\n" + "\n".join(f"• <code>{i}</code>" for i in failed))

    await message.answer("\n\n".join(lines) if lines else "Hech narsa bo'lmadi.")


# ============================================================
# 11b) BROADCAST
# ============================================================
@router.message(Command("broadcast"))
async def cmd_broadcast(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("Bu buyruq faqat admin uchun.")
        return
    admin_broadcast_state[message.from_user.id] = True
    await message.answer(
        "📢 <b>Broadcast rejimi</b>\n\n"
        "Xabaringizni yuboring — matn, rasm, video yoki fayl bo'lishi mumkin.\n"
        "Bekor qilish uchun /cancel"
    )


@router.message(Command("cancel"))
async def cmd_cancel(message: Message):
    if not is_admin(message.from_user.id):
        return
    if admin_broadcast_state.pop(message.from_user.id, False):
        await message.answer("❌ Broadcast bekor qilindi.")
    else:
        await message.answer("Bekor qilinadigan amal yo'q.")


@router.message(F.document)
async def upload_document(message: Message, bot: Bot):
    if not is_admin(message.from_user.id):
        await message.answer("Fayl yuklash faqat admin uchun ruxsat etilgan.")
        return

    # Broadcast rejimida fayl yuborilsa — broadcast sifatida yuboradi
    if admin_broadcast_state.get(message.from_user.id, False):
        admin_broadcast_state.pop(message.from_user.id)
        user_ids = await get_all_approved_user_ids()
        if not user_ids:
            await message.answer("Hozircha tasdiqlangan foydalanuvchilar yo'q.")
            return
        await message.answer(f"⏳ {len(user_ids)} ta foydalanuvchiga yuborilmoqda...")
        success = 0
        failed = 0
        for uid in user_ids:
            if uid in ADMIN_IDS:
                continue
            try:
                await message.copy_to(uid)
                success += 1
            except Exception:
                failed += 1
            await asyncio.sleep(0.05)
        await message.answer(
            f"✅ Yuborildi — <b>{success}</b> ta foydalanuvchi\n"
            f"❌ Yetkazib bo'lmadi — <b>{failed}</b> ta"
        )
        return

    doc: Document = message.document
    file = await bot.get_file(doc.file_id)
    file_bytes_io = await bot.download_file(file.file_path)
    raw_bytes = file_bytes_io.read()

    is_xlsx = (
        doc.file_name and doc.file_name.lower().endswith(".xlsx")
    ) or doc.mime_type in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }

    if is_xlsx:
        await process_approve_xlsx(bot, raw_bytes, message)
        return

    is_txt = (
        doc.mime_type in {"text/plain", "application/octet-stream"}
        or (doc.file_name and doc.file_name.lower().endswith(".txt"))
    )
    if not is_txt:
        await message.answer("Faqat .txt (test) yoki .xlsx (ruxsat ro'yxati) fayl yuboring.")
        return

    content_str = raw_bytes.decode("utf-8-sig")
    try:
        subject, short_name, questions = parse_subject_file(content_str)
    except Exception as e:
        await message.answer(f"❌ Faylda xato bor:\n{e}")
        return
    subject_tests[subject] = questions
    subject_short_names[subject] = short_name
    await save_subject(subject, questions, short_name)
    result_text = (
        f"✅ '<b>{subject}</b>' fani muvaffaqiyatli yuklandi.\n"
        f"Savollar soni: <b>{len(questions)}</b> ta."
    )
    if short_name:
        result_text += f"\nQisqa nom: <b>{short_name}</b>"
    await message.answer(result_text)


# ============================================================
# 12) DEFAULT HANDLER
# ============================================================
@router.message()
async def default_handler(message: Message, bot: Bot):
    if message.chat.type != ChatType.PRIVATE:
        return

    # Broadcast rejimida matn/media xabar kelsa — broadcast sifatida yuboradi
    if is_admin(message.from_user.id) and admin_broadcast_state.get(message.from_user.id, False):
        admin_broadcast_state.pop(message.from_user.id)
        user_ids = await get_all_approved_user_ids()
        if not user_ids:
            await message.answer("Hozircha tasdiqlangan foydalanuvchilar yo'q.")
            return
        await message.answer(f"⏳ {len(user_ids)} ta foydalanuvchiga yuborilmoqda...")
        success = 0
        failed = 0
        for uid in user_ids:
            if uid in ADMIN_IDS:
                continue
            try:
                await message.copy_to(uid)
                success += 1
            except Exception:
                failed += 1
            await asyncio.sleep(0.05)
        await message.answer(
            f"✅ Yuborildi — <b>{success}</b> ta foydalanuvchi\n"
            f"❌ Yetkazib bo'lmadi — <b>{failed}</b> ta"
        )
        return

    await send_welcome(message)


# ============================================================
# 13) RENDER UCHUN HEALTH SERVER
# ============================================================
async def health(request: web.Request) -> web.Response:
    try:
        pool = await get_pool()
        async with pool.acquire() as conn:
            await conn.execute("SELECT 1")
        return web.Response(text="OK - Database Active", status=200)
    except Exception as e:
        # Agar baza ulanishida xato bo'lsa ham 200 qaytaramiz, 
        # toki UptimeRobot Render'ni o'chirib qo'ymasin (Down deb hisoblamasin).
        print(f"Health check database ping error: {e}")
        return web.Response(text="OK - Bot Active (DB Issue)", status=200)


async def start_web_server() -> None:
    app = web.Application()
    app.router.add_get("/", health)
    app.router.add_get("/health", health)
    runner = web.AppRunner(app)
    await runner.setup()
    port = int(os.getenv("PORT", "10000"))
    site = web.TCPSite(runner, host="0.0.0.0", port=port)
    await site.start()


# ============================================================
# 14) ISHGA TUSHIRISH
# ============================================================
async def main():
    global subject_tests, subject_short_names

    try:
        if not BOT_TOKEN:
            raise ValueError("BOT_TOKEN topilmadi. Render Environment ga qo'shing.")
        if not DATABASE_URL:
            raise ValueError("DATABASE_URL topilmadi. Render Environment ga qo'shing.")

        print("Ma'lumotlar bazasiga ulanish kutilmoqda...")
        await init_db()
        subject_tests, subject_short_names = await load_questions()

        bot = Bot(
            token=BOT_TOKEN,
            default=DefaultBotProperties(parse_mode=ParseMode.HTML),
        )
        dp = Dispatcher()
        dp.include_router(router)

        await set_bot_commands(bot)
        await start_web_server()

        print("Web server ishga tushdi...")
        print("Bot ishga tushdi...")
        await dp.start_polling(bot)
    except Exception as e:
        print(f"BOT ISHGA TUSHISHIDA XATOLIK: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    asyncio.run(main())
